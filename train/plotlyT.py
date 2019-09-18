# -*- coding:utf-8 -*-
# import cx_Oracle
#
#
# conn = cx_Oracle.connect("save_omd/omd_save_20170727@10.32.9.14/save")
# cursor = conn.cursor()
# sql = 'select * from order_statistics_xj_view'
# cursor.execute(sql)
# rows = cursor.fetchall()
# print(rows)
# a = ('新疆', 188, 159, '2019-08-23 01:01:01', 68, 'XJ', '国药集团新疆新特药业有限公司', 'XJ7', '国药新疆新特药业伊犁仓', '11-销售', '送货', '公路', '送货',
#      '送货', '907')
import plotly.offline as of
import plotly.graph_objs as go
import chart_studio.plotly as py
import chart_studio.tools as ts
import openpyxl
import cufflinks
import datetime
import numpy as np
import pandas as pd

# ts.set_config_file(world_readable=True, sharing='public')     # online
cufflinks.go_offline(connected=True)            # offline
of.offline.init_notebook_mode(connected=True)   # offline
start_time = datetime.datetime.now()
print('Start time：', start_time)
# wb = openpyxl.load_workbook("D:\\SoftWare\\BaiduYun\\BaiduNetdiskDownload\\HB\\HB-custkssx.xlsx")
wb = openpyxl.load_workbook("D:\\Work\\MyProject\\Pub\\train1.xlsx")
# ws = wb['Sheet1']
ws = wb['train']
cell_new = 0
live = 0
die = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    for i in row:
        if i.value == 0:
            die += 1
        if i.value == 1:
            live += 1
print('survived:', live)
print('not survived:', die)
colors = ['#96D38C', '#FEBFB3', '#E1396C', '#D0F9B1']
col_y = [live, die]
col_x = [1, 0]
trace0 = go.Bar(
    x=col_x,
    y=col_y,
    # marker=dict(colors=colors, line=dict(color='#000000', width=2))
)
data = [trace0]
of.plot(data, filename='survived.html')     # py.plot会生成一个离线的html文件,里面放置图片。而py.iplot则直接在ipython notebook里面生成图片
### 折线
# col_x = []
# col_y = []
#### iter_rows ws.rows ws.columns
# for i in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1):
#     for col in i:
#         col_x.append(col.value)
# for i in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=3):
#     for col1 in i:
#         if col1.value == '客户':
#             cell_new = 1
#         if col1.value == '供应商':
#             cell_new = 2
#         if col1.value == '客商':
#             cell_new = 3
#         col_y.append(cell_new)

# trace0 = go.Scatter(
#     x=col_x,
#     y=col_y,
#     mode='markers'
# )

# data = [trace0]
# wb.close()
# of.plot({"data": data, "layout": go.Layout(title='HB客商分散图')}) # plot


### 饼图
# labels = []
# values = []
# colors = ['#96D38C', '#FEBFB3', '#E1396C', '#D0F9B1']
# # colors = ['#FEBFB3', '#E1396C', '#96D38C', '#D0F9B1']
# for i in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2, min_col=2):
#     for rows1 in i:
#         labels.append(rows1.value)
# for i in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=3, min_col=3):
#     for rows2 in i:
#         values.append(rows2.value)
# trace = go.Pie(labels=labels,
#                values=values,
#                hoverinfo='label+percent',
#                textinfo='value',
#                textfont=dict(size=15),
#                marker=dict(colors=colors, line=dict(color='#000000', width=2))
#                )
# data = [trace]
# # of.plot(data, filename='hb_customer_code.html')
# py.iplot({"data": data, "layout": go.Layout(title='HB-Attribute distribution')}, filename='HBks', auto_open=True)
# stop_time = datetime.datetime.now()
# print('Stop time：', stop_time)
# print('Time consuming：', stop_time - start_time)
