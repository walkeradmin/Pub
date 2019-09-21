# -*- coding:utf-8 -*-
# author： walker

import cx_Oracle
import win32gui
import win32api
from win32.lib import win32con
import win32clipboard as w
import time
import datetime
import shelve
import os
import threading
from threading import Lock   # Thread Timer
import ctypes
from ctypes import wintypes
import pythoncom
import logformat
import requests
import json
import configparser
import ast
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font    # PatternFill
import sys
import re
import schedule
from selenium import webdriver
from selenium.webdriver import ActionChains
import pymysql


def remove_bom(config_path):  # BOM字节
    content = open(config_path).read()
    content = re.sub(r"\xfe\xff", "", content)
    content = re.sub(r"\xff\xfe", "", content)
    content = re.sub(r"\xef\xbb\xbf", "", content)
    open(config_path, 'w').write(content)


def deploy():
    conf_file = os.path.join(os.getcwd(), 'DevopsConf.ini')
    cp = configparser.ConfigParser()
    cp.read(conf_file)
    return cp


class ORDER:
    def __init__(self):
        self._fre, self._flag1, self._ia = 0, 0, 0
        self._or_tem_mess = deploy().get('ins', 'order_mess')
        self._method_mess = deploy().get('ins', 'method_mess')
        self._allTime = eval(deploy().get('Timer', 'allTime'))
        self._folderAll = os.getcwd() + "\\formatLogs\\allOrderFile\\"
        self._a, self._b, self._c, self._d, self._e, self._f = -1, -2, -3, -4, -5, -6
        self._sql_all_1 = deploy().get('SQL', 'order_all')
        self._sql_ma_1 = deploy().get('SQL', 'order_ma')
        self._info2 = 'CLIP_FILE task perform succeed'
        self._info3 = 'Running send file task'
        self._info4 = 'send file success'

    def timer_all(self):
        now = time.strftime("%H:%M:%S")
        week = time.strftime("%a")
        clock_order = eval(deploy().get('Clock', 'clock_order'))
        if clock_order[0] <= now <= clock_order[1] and week != 'Sat' and week != 'Sun':
            self.query()
        else:
            log().info(
                '| ORDER CLASS | TIMER FUNCTION | Message：'
                'Time to stop pushing, the next query will start at 08:00 tomorrow morning |')

    def start_time(self):
        global clock
        week = time.strftime("%A")
        if week == 'Monday':
            now_time = datetime.datetime.now()
            time_format = now_time.strftime('%Y-%m-%d')
            clock = time_format + ' ' + '00:00:00'
        elif week == 'Tuesday':
            now_time = datetime.datetime.now()
            change_time = now_time + datetime.timedelta(days=self._a)
            time_format = change_time.strftime('%Y-%m-%d')
            clock = time_format + ' ' + '00:00:00'
        elif week == 'Wednesday':
            now_time = datetime.datetime.now()
            change_time = now_time + datetime.timedelta(days=self._b)
            time_format = change_time.strftime('%Y-%m-%d')
            clock = time_format + ' ' + '00:00:00'
        elif week == 'Thursday':
            now_time = datetime.datetime.now()
            change_time = now_time + datetime.timedelta(days=self._c)
            time_format = change_time.strftime('%Y-%m-%d')
            clock = time_format + ' ' + '00:00:00'
        elif week == 'Friday':
            now_time = datetime.datetime.now()
            change_time = now_time + datetime.timedelta(days=self._d)
            time_format = change_time.strftime('%Y-%m-%d')
            clock = time_format + ' ' + '00:00:00'
        elif week == 'Saturday':
            now_time = datetime.datetime.now()
            change_time = now_time + datetime.timedelta(days=self._e)
            time_format = change_time.strftime('%Y-%m-%d')
            clock = time_format + ' ' + '00:00:00'
        elif week == 'Sunday':
            now_time = datetime.datetime.now()
            change_time = now_time + datetime.timedelta(days=self._f)
            time_format = change_time.strftime('%Y-%m-%d')
            clock = time_format + ' ' + '00:00:00'
        return clock

    def send(self, send_mess, group_name):
        try:
            mutex.acquire()
            w.OpenClipboard()
            w.EmptyClipboard()
            if type(send_mess) == str:
                w.SetClipboardData(win32con.CF_UNICODETEXT, send_mess)
                w.CloseClipboard()
                # time wait pywintypes.error: (1418, 'GetClipboardData',线程没有打开的剪贴板)
                time.sleep(1)
                handle = win32gui.FindWindow('TXGuiFoundation', group_name)
                win32gui.SendMessage(handle, 770, 0, 0)
                win32gui.SendMessage(handle, win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)
                log().info('| ORDER CLASS | SEND MESS FUNCTION | Message：Group {} successfully sent the task |'.format(
                    group_name))
            if type(send_mess) == bytes:
                w.SetClipboardData(w.CF_HDROP, send_mess)
                log().info('| ORDER CLASS | SEND MESS FUNCTION | Message：' + self._info2 + ' |')
                w.CloseClipboard()
                log().info('| ORDER CLASS | SEND MESS FUNCTION | Message：' + self._info3 + ' |')
                win = win32gui.FindWindow(None, group_name)
                time.sleep(1)
                win32api.PostMessage(win, win32con.WM_PASTE, 0, 0)
                time.sleep(1)
                win32gui.SendMessage(win, win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)
                log().info('| ORDER CLASS | SEND MESS FUNCTION | Message：' + self._info4 + ' |')
            mutex.release()
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())

    def format_mess_all(self, result, group_name):
        try:
            met = eval(deploy().get('Method', 'dic1'))
            num = "".join(result[:1])       # index + 1
            cause = "".join(result[1:2])
            or_id = "".join(result[2:3])
            ow = "".join(result[3:4])
            depot = "".join(result[4:5])
            or_ty = "".join(result[5:6])
            or_time = "".join(result[7:8])
            global method
            if any(i in cause for i in met.keys()):
                for k, v in met.items():
                    if k in cause:
                        method = v
            else:
                method = self._method_mess
            send_mess = self._or_tem_mess.format(num, cause, or_id, ow, depot, or_ty, or_time, method)
            self.send(send_mess, group_name)
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())

    def query(self):
        try:
            self._fre += 1
            dic = ast.literal_eval(deploy().get('Group', 'dic'))
            push_time = time.strftime("%Y-%m-%d %H:%M:%S %A")
            day = time.strftime("%Y-%m-%d 00:00:00")
            log().info('| ORDER CLASS | QUERY FUNCTION | Message：Sql query time：' + self.start_time() + ' |')
            log().info('| ORDER CLASS | QUERY FUNCTION | Message：Start pushing：' + push_time + ' |')
            log().info('| ORDER CLASS | QUERY FUNCTION | Message：Start query for the：' + str(self._fre) + ' time |')
            obj5 = Oracle()
            rows_all = obj5.execute_sql(self._sql_all_1.format(self.start_time(), self.start_time(), self.start_time()))
            rows_ma = obj5.execute_sql(self._sql_ma_1.format(day))
            obj5.close()
            rows_all.extend(rows_ma)
            while True:
                if self._flag1 == 0:
                    for k, v in dic.items():
                        count1, count2, count3 = 0, 0, 0
                        for i in rows_all:
                            tuple1 = i
                            if k in tuple1[7] \
                                    and v[0] == '0' \
                                    and v[1] == '1' \
                                    and v[2] == '1' \
                                    and '委托入库' not in tuple1[5] \
                                    and '物流调整' not in tuple1[5]:  # only out order
                                count1 += 1
                                order1 = list(tuple1)
                                order1.insert(0, str(count1))
                                result1 = tuple(order1)
                                log().info('| ORDER CLASS | QUERY FUNCTION | Message：' + k + '卡单 |')
                                log().info('| ORDER CLASS | QUERY FUNCTION | Message：' + str(result1) + ' |')
                                self.format_mess_all(result1, v[3])
                            if k in tuple1[7] \
                                    and v[0] == '0' \
                                    and v[1] == '0' \
                                    and v[2] == '1' \
                                    and '物流调整' not in tuple1[5]:  # oder out in
                                count2 += 1
                                order2 = list(tuple1)
                                order2.insert(0, str(count2))
                                result2 = tuple(order2)
                                log().info('| ORDER CLASS | QUERY FUNCTION | Message：' + k + '卡单 |')
                                log().info('| ORDER CLASS | QUERY FUNCTION | Message：' + str(result2) + ' |')
                                self.format_mess_all(result2, v[3])
                            if k in tuple1[7] \
                                    and v[0] == '0' \
                                    and v[1] == '0' \
                                    and v[2] == '0':  # out in adj
                                count3 += 1
                                order3 = list(tuple1)
                                order3.insert(0, str(count3))
                                result3 = tuple(order3)
                                log().info('| ORDER CLASS | QUERY FUNCTION | Message：' + k + '卡单 |')
                                log().info('| ORDER CLASS | QUERY FUNCTION | Message：' + str(result3) + ' |')
                                self.format_mess_all(result3, v[3])
                    self._flag1 = 1
                else:
                    log().info(
                        '| ORDER CLASS | QUERY FUNCTION | Message：Order：Latest data volume：' + str(
                            len(rows_all)) + ' |')
                    log().info('| ORDER CLASS | QUERY FUNCTION | Message：Order：Number of historical data：' + str(
                        len(his_rows_all)) + ' |')
                    if not os.path.exists(self._folderAll):
                        os.makedirs(self._folderAll)
                    day = time.strftime("%Y-%m-%d")
                    f = open(self._folderAll + 'log_all_%s.txt' % day, 'w+', encoding='utf-8')
                    f.write(str(his_rows_all))
                    f.close()
                    check = list(set(tuple(rows_all)).difference(set(tuple(his_rows_all))))
                    if check:
                        his_rows_all.extend(check)
                        rows_all = check
                        self._flag1 = 0
                    else:
                        self._ia += 1
                        if self._ia == 1:  # first query
                            break
                        else:
                            log().info(
                                '| ORDER CLASS | QUERY FUNCTION | Message：'
                                'Order The query results are the same as the previous result,'
                                'eliminating the need to perform a send task and '
                                'preparing for a new round of queries |')
                            break
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())

    def query_his(self):
        global his_rows_all
        global his_rows_ma
        try:
            obj4 = Oracle()
            his_rows_all = obj4.execute_sql(
                self._sql_all_1.format(self.start_time(), self.start_time(), self.start_time()))
            his_rows_ma = obj4.execute_sql(self._sql_ma_1.format(self.start_time()))
            his_rows_all.extend(his_rows_ma)    # merge
            obj4.close()
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())


class DROP(ctypes.Structure):
    _fields_ = (('pFiles', wintypes.DWORD),
                ('pt',     wintypes.POINT),
                ('fNC',    wintypes.BOOL),
                ('fWide',  wintypes.BOOL))


class FILE:
    def __init__(self):
        self._sf_jd_mess = deploy().get('ins', 'sf_jd_mess')
        self._ef_jd_mess = deploy().get('ins', 'ef_jd_mess')
        self._folderKh = os.getcwd() + "\\formatLogs\\SanFangOutInFile\\"
        self._folder_out = os.getcwd() + "\\formatLogs\\ErFangOutFile\\"
        self._khTime = eval(deploy().get('Timer', 'khTime'))
        self._info1 = 'Copying to clipboard, filename：'
        self._sql_all_rec = deploy().get('SQL', 'sql_all_rec')
        self._sql_all_feed = deploy().get('SQL', 'sql_all_feed')
        self._title = eval(deploy().get('excel', 'title'))
        self._title1 = eval(deploy().get('excel', 'title1'))
        self._code = 'utf-8'
        self._letter1 = 'Z'
        self._letter2 = 'A'

    def clip_files(self, file_list, group_name):
        try:
            offset = ctypes.sizeof(DROP)
            length = sum(len(p) + 1 for p in file_list) + 1
            size = offset + length * ctypes.sizeof(ctypes.c_wchar)
            buf = (ctypes.c_char * size)()
            df = DROP.from_buffer(buf)
            df.pFiles, df.fWide = offset, True
            for path in file_list:
                log().info('| FILE CLASS | CLIP FILES FUNCTION | Message：' + self._info1 + path + ' |')
                array_t = ctypes.c_wchar * (len(path) + 1)
                path_buf = array_t.from_buffer(buf, offset)
                path_buf.value = path
                offset += ctypes.sizeof(path_buf)
            stg = pythoncom.STGMEDIUM()
            stg.set(pythoncom.TYMED_HGLOBAL, buf)
            test = ORDER()
            test.send(stg.data, group_name)
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())

    def format_mess_file(self, rows_rec, rows_feed):
        try:
            owner_name = deploy().items('Stas')
            for na in owner_name:
                ow_name = na[0]
                if rows_rec:
                    rec_out = list()
                    rec_in = list()
                    for i3 in rows_rec:
                        if i3[0] == ow_name and i3[2] == '入库':
                            tup_in_rec = i3
                            rec_in.extend(tup_in_rec[1:2])
                    rec_num_in = len(rec_in)
                    for i1 in rows_rec:
                        if i1[0] == ow_name and i1[2] == '出库':
                            tup_out_rec = i1
                            rec_out.extend(tup_out_rec[1:2])
                    rec_num_out = len(rec_out)
                else:
                    rec_in, rec_out = [], []
                    rec_num_in = len(rows_rec)
                    rec_num_out = len(rows_rec)
                if rows_feed:
                    feed_in = list()
                    feed_out = list()
                    for i4 in rows_feed:
                        if i4[0] == ow_name and i4[2] == '入库':
                            tup_in_feed = i4
                            feed_in.extend(tup_in_feed[1:2])
                    feed_num_in = len(feed_in)
                    for i2 in rows_feed:
                        if i2[0] == ow_name and i2[2] == '出库':
                            tup_out_feed = i2
                            feed_out.extend(tup_out_feed[1:2])
                            # filename = tup_out_feed[2:3]
                    feed_num_out = len(feed_out)
                else:
                    feed_in, feed_out = [], []
                    feed_num_in = len(rows_feed)
                    feed_num_out = len(rows_feed)
                s_time = time.strftime("%Y-%m-%d %H:%M:%S")
                send_mess = self._sf_jd_mess.format(ow_name, s_time, str(rec_num_in), str(rec_num_out),
                                                    str(feed_num_in), str(feed_num_out))
                if not os.path.exists(self._folderKh):
                    os.makedirs(self._folderKh)
                group_li = eval(deploy().get('Stas', ow_name))
                day = time.strftime("%Y-%m-%d")
                wb = Workbook()
                ws = wb.active
                ws.title = '截单单统计'
                cell_font = Font(name='Microsoft YaHei UI', size=11, bold=True)  # italic 斜体 bold 加粗
                order_font = Font(name='Consolas')
                # cell_fill = PatternFill("solid", fgColor="5CACEE")
                ws.append(self._title)
                # ws.iter_rows() all cell
                col_max = len(self._title)
                letter = self.num_column(col_max) + '1'
                for row in ws["A1":letter]:  # <Cell 'sheet'.A1>
                    for eachCell in row:
                        eachCell.font = cell_font
                        # eachCell.fill = cell_fill
                ws.move_range("A1", rows=0)
                ws.append([ow_name, rec_num_in, rec_num_out])
                ws['F2'] = feed_num_in
                ws['G2'] = feed_num_out
                ws['J2'] = time.strftime("%Y-%m-%d %H:%M:%S")
                ws['A2'].font = cell_font
                ws['B2'].font = order_font
                ws['C2'].font = order_font
                ws['F2'].font = order_font
                ws['G2'].font = order_font
                ws['J2'].font = order_font
                for i in range(1, ws.max_column+1):
                    cell_col = self.num_column(i)
                    ws.column_dimensions[cell_col].width = 20        # modify cell wide
                for i1 in range(len(rec_in)):
                    ws.cell(i1 + 2, 4, value=rec_in[i1]).font = order_font
                for i2 in range(len(rec_out)):
                    ws.cell(i2 + 2, 5, value=rec_out[i2]).font = order_font
                for i3 in range(len(feed_in)):
                    ws.cell(i3 + 2, 8, value=feed_in[i3]).font = order_font
                for i4 in range(len(feed_out)):
                    ws.cell(i4 + 2, 9, value=feed_out[i4]).font = order_font
                if group_li[0] == 0:
                    wb.save(self._folderKh + "%s_%s.xlsx" % (ow_name, day))
                    wb.close()
                    send = ORDER()
                    send.send(send_mess, group_li[1])
                    day = time.strftime("%Y-%m-%d")
                    file_name = self._folderKh + '%s_%s.xlsx' % (ow_name, day)
                    check = []
                    if rows_rec == check and rows_feed == check:
                        pass
                    else:
                        self.clip_files([os.path.abspath(file_name)], group_li[1])
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())

    def all_mess_file(self, result, cen_code):
        try:
            center_name = deploy().items('Center')
            if not os.path.exists(self._folder_out):
                os.makedirs(self._folder_out)
            for cen in center_name:
                dtl_sum = 0
                hdr_sum = 0
                cen_name = cen[0]
                li_center = []
                group_li = eval(deploy().get('Center', cen_name))
                if result:
                    for row_num1 in range(len(result)):
                        li_center.append(result[row_num1][0])
                    if any(i in cen_name for i in li_center):   # judge center in or not in rows
                        for tup in result:
                            if tup[0] == cen_name:              # same center sum
                                dtl_sum = dtl_sum + int(tup[1])
                                hdr_sum = hdr_sum + int(tup[2])
                        s_time = time.strftime("%Y-%m-%d %H:%M:%S")
                        mess = self._ef_jd_mess.format(cen_name, s_time, str(dtl_sum), str(hdr_sum))
                        if group_li[0] == 0:
                            send = ORDER()
                            send.send(mess, group_li[1])
                            day = time.strftime("%Y-%m-%d")
                            wb = Workbook()
                            ws = wb.active
                            ws.title = '截单统计'
                            cell_font = Font(name='Microsoft YaHei UI', size=11, bold=True)
                            order_font = Font(name='Consolas')
                            ws.append(self._title1)
                            for i in range(1, ws.max_column+1):
                                cell_col = self.num_column(i)
                                ws.column_dimensions[cell_col].width = 27        # modify cell wide
                            col_max = len(self._title1)
                            letter = self.num_column(col_max) + '1'
                            for row in ws["A1":letter]:  # <Cell 'sheet'.A1>
                                for eachCell in row:
                                    eachCell.font = cell_font
                            ws.move_range("A1", rows=0)
                            fields_num = len(result[0])
                            for row_num in range(len(result)):
                                if cen_name == result[row_num][0]:
                                    for i1 in range(fields_num):
                                        ws.cell(row_num+2, i1+1, value=result[row_num][i1]).font = order_font
                            wb.save(self._folder_out + "%s_%s.xlsx" % (cen_name, day))
                            wb.close()
                            file_name = self._folder_out + '%s_%s.xlsx' % (cen_name, day)
                            self.clip_files([os.path.abspath(file_name)], group_li[1])
                else:
                    log().info('| FILE CLASS | ALL MESS FILE FUNCTION |Message：' + cen_code + '：No query results |')
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())

    # statistics column name
    def num_column(self, num):
        interval = ord(self._letter1) - ord(self._letter2)
        tmp = ''
        multiple = num // interval
        remainder = num % interval
        while multiple > 0:
            if multiple > 25:
                tmp += 'A'
            else:
                tmp += chr(64 + multiple)
            multiple = multiple // interval
        tmp += chr(64 + remainder)
        return tmp

    # statistics value size not use
    def len_byte(self, value):
        if value is None or value == "":
            return 10
        if type(value) != int:
            length = len(value)
            utf8_length = len(value.encode(self._code))
            length = (utf8_length - length) / 2 + length
        else:
            length = len(str(value))
        return int(length)

    def query_kh(self):
        obj4 = Oracle()
        rows_rec = obj4.execute_sql(self._sql_all_rec)
        rows_feed = obj4.execute_sql(self._sql_all_feed)
        self.format_mess_file(rows_rec, rows_feed)

    def query_all_out(self, center_code):
        sql_ef_out = deploy().get('SQL', 'sql_xj')
        obj41 = Oracle()
        rows_all_out = obj41.execute_sql(sql_ef_out.format(center_code))
        self.all_mess_file(rows_all_out, center_code)

    def timer_kh(self):
        date = time.strftime("%A")
        now = datetime.datetime.now()
        clock_kh = eval(deploy().get('Clock', 'clock_kh'))
        clock_ef_out = eval(deploy().get('Clock', 'clock_all_out'))
        if date == 'Saturday' or date == 'Sunday':
            log().info(
                '| FILE CLASS | TIMER KH FUNCTION | Message：'
                '[SanFangOutInFile]-[ErFangOutFile] Do not perform tasks on weekends |')
        else:
            for i in clock_kh:
                send_time = time.strftime("%Y-%m-%d {}".format(i))
                of_time = datetime.datetime.strptime(send_time, "%Y-%m-%d %H:%M:%S")
                # 11:30:00.xx < x < 11:31:00.xx
                # 16:00:00.xx < x < 16:01:00.xx
                if of_time < now < of_time + datetime.timedelta(minutes=1):
                    self.query_kh()
                else:
                    log().info(
                        '| FILE CLASS | TIMER KH FUNCTION | Message：'
                        'KH timer {} Do not perform tasks during non-working hours |'.format(i))

            for dic1 in clock_ef_out:
                i1 = clock_ef_out[dic1]
                send_time1 = time.strftime("%Y-%m-%d {}".format(i1))
                of_time1 = datetime.datetime.strptime(send_time1, "%Y-%m-%d %H:%M:%S")
                if of_time1 < now < of_time1 + datetime.timedelta(minutes=1):
                    self.query_all_out(dic1)
                else:
                    log().info(
                        '| FILE CLASS | TIMER KH FUNCTION | Message：'
                        'All ef owner center {} timer {} Do not perform tasks during non-working hours |'.format(
                            dic1, i1))


class SH:
    def __init__(self):
        self._sh_sn_mess = deploy().get('ins', 'sh_sn_mess')
        self._erp_mess = deploy().get('ins', 'erp_mess')
        self._ots1 = deploy().get('ins', 'ots1')
        self._ots2 = deploy().get('ins', 'ots2')
        self._erpTime = eval(deploy().get('Timer', 'erpTime'))
        self._hrTime = eval(deploy().get('Timer', 'hrTime'))
        self._folderErp = os.getcwd() + "\\formatLogs\\erpLog\\"
        self._mess = '辉瑞卡单，请及时处理'
        self._sql_sh_sn = deploy().get('SQL', 'sql_sh_sn')
        self._name = eval(deploy().get('Name', '上海SAVE运维主管'))
        self._gy_group_name = eval(deploy().get('Sinopharm', 'GYWL'))

    def timer_erp(self):
        date = time.strftime("%A")  # %a
        clock_erp = eval(deploy().get('Clock', 'clock_erp'))
        if date == 'Saturday' or date == 'Sunday':
            log().info('| SH CLASS | TIMER ERP FUNCTION | Message：ERP task do not perform tasks on weekends |')
        else:
            now = datetime.datetime.now()
            for i in clock_erp:
                send_time = time.strftime("%Y-%m-%d " + i)
                range_time = datetime.datetime.strptime(send_time, "%Y-%m-%d %H:%M:%S")
                # 12:00:00.xx < x < 12:01:00.xx
                # 08:30:00.xx < x < 08:31:00.xx
                if range_time < now < range_time + datetime.timedelta(minutes=1):
                    self.erp(i)
                else:
                    log().info('| SH CLASS | TIMER ERP FUNCTION | Message：ERP task do not perform send task |')

    def timer_hr(self):
        range_time1 = datetime.datetime.now()
        date = time.strftime("%A")  # %a
        if date == 'Saturday' or date == 'Sunday':
            log().info('| SH CLASS | TIMER HR FUNCTION | Message：HR Do not perform tasks on weekends |')
        else:
            now = datetime.datetime.now()
            if range_time1 <= now < range_time1 + datetime.timedelta(seconds=1):
                self.check_feed()

    def format_sh_sn(self, sh_num, sn_num):
        try:
            tii = time.strftime("%Y-%m-%d %H:%M:%S")
            mess = self._sh_sn_mess.format(tii, str(sh_num), str(sn_num))
            send = ORDER()
            if self._gy_group_name[0] == 0:
                send.send(mess, self._gy_group_name[1])
            log().info(
                '| SH CLASS | FORMAT SH SN FUNCTION | Message：上海出库 ' + str(sh_num) + ' 枢纽出库 ' + str(sn_num) + ' |')
            if not os.path.exists(self._folderErp):
                os.makedirs(self._folderErp)
            file_now = time.strftime("%Y-%m-%d")
            f = open(self._folderErp + 'Check_Order%s.txt' % file_now, 'w+', encoding='utf-8')
            f.write(mess)
            f.close()
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())

    def format_erp(self, result, clo):
        try:
            name_li = deploy().items('Name')
            erp_quantity = result[0]
            erp_in = erp_quantity[1]
            erp_in_del = erp_quantity[2]
            erp_out = erp_quantity[3]
            erp_out_del = erp_quantity[4]
            save_quantity = result[1]
            save_in = save_quantity[1]
            save_in_del = save_quantity[2]
            save_out = save_quantity[3]
            save_out_del = save_quantity[4]
            now = time.strftime("%Y-%m-%d %H:%M:%S")
            mess = self._erp_mess.format(now, str(erp_in), str(erp_in_del), str(erp_out), str(erp_out_del),
                                         str(save_in), str(save_in_del), str(save_out), str(save_out_del))
            zw = eval(deploy().get('Clock', 'clock_erp'))[0]
            if erp_in == save_in and erp_in_del == save_in_del and erp_out == save_out and erp_out_del == save_out_del:
                mess0 = mess + 'Query result：ERP == SAVE'
                log().info('| SH CLASS | FORMAT ERP FUNCTION | Message: ' + mess0 + ' |')
                if not os.path.exists(self._folderErp):
                    os.makedirs(self._folderErp)
                file_now = time.strftime("%Y-%m-%d")
                f = open(self._folderErp + 'ERP_SH_SN%s.txt' % file_now, 'w+', encoding='utf-8')
                f.write(mess0)
                f.close()
                send1 = ORDER()
                if clo == zw:
                    for i in name_li:
                        name_list = eval(i[1])[1]
                        if eval(i[1])[0] == 0:
                            send1.send(mess0, name_list)
                else:
                    self.order_sh_sn()
            else:
                mess1 = mess + '''
Query result：ERP != SAVE''' + self._ots1
                meo = self._ots2
                log().info(
                    '| SH CLASS | FORMAT ERP FUNCTION | Message: '
                    'Query result ERP != SAVE，Send data results to 上海SAVE运维主管 |')
                log().info('| SH CLASS | FORMAT ERP FUNCTION | Message：' + mess + ' |')
                send2 = ORDER()
                for i in name_li:
                    name_list = eval(i[1])[1]
                    if eval(i[1])[0] == 0:              # judge send or not send message
                        mes = mess1.format(name_list)
                        send2.send(mes, name_list)     # 上海SAVE运维主管、全国SAVE运维、other person
                if clo != zw:
                    if self._gy_group_name[0] == 0:
                        send2.send(meo, self._gy_group_name[1])  # 国药物流SAVE+WMS+TMS
                if not os.path.exists(self._folderErp):
                    os.makedirs(self._folderErp)
                file_now = time.strftime("%Y-%m-%d")
                f = open(self._folderErp + 'ERP_SH_SN%s.txt' % file_now, 'w+', encoding='utf-8')
                f.write(mess)
                f.close()
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())

    def format_hr(self, feed_rows):
        try:
            if feed_rows:
                if self._name[0] == 0:
                    send = ORDER()
                    send.send(self._mess, self._name[1])   # 上海SAVE运维主管
                    log().info(
                        '| SH CLASS | FORMAT HR FUNCTION | Message：' + str(
                            feed_rows) + ', HR Query the data, please handle it in time,re-query after one hour |')
            else:
                log().info('| SH CLASS | FORMAT HR FUNCTION | Message：HR No data was found, re-query after one hour |')
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())

    def order_sh_sn(self):
        obj3 = Oracle()
        sh_sn_num = obj3.execute_sql(self._sql_sh_sn)
        sh_num = (sh_sn_num[0])[1]
        sn_num = (sh_sn_num[1])[1]
        obj3.close()
        self.format_sh_sn(sh_num, sn_num)

    def erp(self, clo):
        db_link = deploy().get('SQL', 'db_link')
        sql_erp = deploy().get('SQL', 'sql_erp')
        obj2 = Oracle()
        obj2.execute_sql(db_link)
        erp_rows = obj2.execute_sql(sql_erp)
        obj2.close()
        self.format_erp(erp_rows, clo)

    def check_feed(self):
        sql_hr_feed = deploy().get('SQL', 'sql_feed')
        obj1 = Oracle()
        feed_rows = obj1.execute_sql(sql_hr_feed)
        obj1.close()
        self.format_hr(feed_rows)


class Oracle:
    def __init__(self):
        self.user = user
        self.password = password
        self.ip = ip
        self.name = name
        self.conn = None
        self.cursor = None
        self.isClose = True

    def connect(self):          # connect
        try:
            self.conn = cx_Oracle.connect(self.user+'/'+self.password+'@'+self.ip+'/'+self.name)
            self.cursor = self.conn.cursor()
            self.isClose = False
        except Exception as e:
            error().error("| ORACLE CLASS | CONNECT FUNCTION | Message："
                          "Oracle connect error![MSG={},USER={},PASSWORD={},IP={},NAME={}] |".format(str(e), self.user,
                                                                                                     self.password,
                                                                                                     self.ip,
                                                                                                     self.name))
            obj1 = Comp('@all', traceback.format_exc())
            obj1.post()
            self.cursor = None
            self.isClose = True
        return self.cursor

    def is_connect(self):           # connected or not
        if self.cursor is not None:
            return True
        else:
            return False

    def close(self):                # close oracle connect
        self.cursor.close()
        self.conn.close()
        self.isClose = True

    def execute_sql(self, sql):     # execute fetchall
        if not self.is_connect():
            self.connect()          # True = Not False
            num = 0
            for i in range(3):      # judge 3 time
                num += 1
                if not self.is_connect():
                    self.connect()
                    log().info(
                        '| ORACLE CLASS | EXECUTE SQL FUNCTION | Message：'
                        'Try to reconnect to the database for the {} time |'.format(
                            str(num)))
                    time.sleep(1)
                else:
                    break
        else:                       # False = Not True
            pass
        try:
            self.cursor.execute(sql)
            if 'declare begin' in sql:
                pass
            else:
                global sql_rows
                sql_rows = self.cursor.fetchall()
        except Exception as e:
            error().error("| ORACLE CLASS | EXECUTE SQL FUNCTION | Message："
                          "Execute sql error![MSG={},SQL={}] |".format(str(e), sql))
            obj1 = Comp('@all', traceback.format_exc())
            obj1.post()
            return ''
        else:
            return sql_rows


class Comp(object):
    def __init__(self, tor, message):
        self._comp_id = ""
        self._secret = ""
        self._agent_id = ""
        self._token = 'https://qyapi.weixin.qq.com/cgi-bin/gettoken?corpid={}&corpsecret={}'.format(self._comp_id,
                                                                                                    self._secret)
        self._tor = tor
        self._message = message

    def conn(self):
        try:
            request = requests.get(self._token)
            access_token = request.json()['access_token']
        except Exception as e:
            log().error('| Comp Class-Function conn | Message：' + str(e))
        else:
            global msg_url
            msg_url = 'https://qyapi.weixin.qq.com/cgi-bin/message/send?access_token={}'.format(access_token)
        return msg_url

    def format(self):

        params = {"touser": self._tor,
                  # "toparty": toparty,
                  "msgtype": "text",
                  "agentid": self._agent_id,
                  "text": {"content": self._message},
                  "safe": 0
                  }
        return params

    def post(self):
        try:
            requests.post(self.conn(), data=json.dumps(self.format()))
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())
        else:
            log().info(
                '| COMP CLASS | POST FUNCTION | Message：'
                'Sending a successful company WeChat | ' + 'sendto：' + self._tor + ';;mess：' + self._message)


class CHECK:
    def __init__(self):
        self._check_c = 0

    def check_thread(self):     # Concurrency thread check
        try:
            self._check_c += 1
            thread_enu = threading.enumerate()
            thread_count = threading.active_count()
            log().info('| CHECK THREAD FUNCTION | Message：{}'.format(thread_enu))
            log().info('| CHECK THREAD FUNCTION | Message：Check Thread alive count {}'.format(thread_count))
            mess = '''Error：{}
Mess：Thread stopped abnormally
Current：number of threads {}
Actually：number of threads 1
Time：{}'''
            if thread_count < 1 and self._check_c != 1:
                log().error('The program thread has an exception, now the number of threads is {}'.format(thread_count))
                mess1 = mess.format(py_name, thread_count, datetime.datetime.now())
                obj1 = Comp('@all', mess1)
                obj1.post()
                # restore(thread_count)
            else:
                if thread_count > 1:
                    mess2 = mess.format(thread_count, datetime.datetime.now())
                    obj2 = Comp('@all', mess2)
                    obj2.post()
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())


def clean_screen():
    os.system('cls')
    log().info('| ClEAN SCREEN FUNCTION | Message：Clear screen task execution completed |')


class ALARM:
    def __init__(self):
        self._username = 'username'
        self._password = 'password'
        self._login_elt = '//*[@id="login-view"]/form/div[3]/button'
        self._mq_elt = '//*[@id="panel-38"]'
        self._gra = 'Grafana'
        self._res = 'pic'
        self._alarm_sql = 'SELECT * FROM zabbix.mq_fe_receive_view'

    def alarm_send(self, result, group_name):
        handle = win32gui.FindWindow(None, group_name)
        if result == self._res:
            w.OpenClipboard()
            w.EmptyClipboard()
            w.CloseClipboard()
            win32api.keybd_event(0x91, 0, 0, 0)  # 0x91 --> win key
            win32api.keybd_event(0x2C, 0, 0, 0)  # 0x2C --> PRINT SCREEN key
            win32api.keybd_event(0x91, 0, win32con.KEYEVENTF_KEYUP, 0)
            win32api.keybd_event(0x2C, 0, win32con.KEYEVENTF_KEYUP, 0)
            time.sleep(1)
            win32gui.SendMessage(handle, 770, 0, 0)
            win32gui.SendMessage(handle, win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)
        elif result != self._res:
            w.OpenClipboard()
            w.EmptyClipboard()
            w.SetClipboardData(win32con.CF_UNICODETEXT, str(result))
            w.CloseClipboard()
            time.sleep(1)
            win32gui.SendMessage(handle, 770, 0, 0)
            if result.startswith('@'):
                time.sleep(1)
                win32gui.SendMessage(handle, win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)
            else:
                win32gui.SendMessage(handle, win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)

    def element(self, url, user_mq, password_mq, group):
        try:
            drive = webdriver.Chrome(".\\chromedriver.exe")  # browser
            drive.maximize_window()  # my window size: {'width': 1514, 'height': 974}!!!
            drive.get(url)
            drive.implicitly_wait(10)
            drive.find_element_by_name(self._username).send_keys(user_mq)
            drive.find_element_by_name(self._password).send_keys(password_mq)
            drive.find_element_by_xpath(self._login_elt).click()  # login
            drive.switch_to.window(drive.window_handles[0])  # Crawl the current page element
            element1 = drive.find_element_by_xpath(self._mq_elt)
            ActionChains(drive).move_to_element(element1).perform()
            time.sleep(3)
            self.alarm_send(self._res, group)
            log().info('| ALARM CLASS | ELEMENT FUNCTION | Message：Screenshot task completed |')
            drive.quit()
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())

    def format_mess(self, rows):
        try:
            if not os.path.exists(alarm_path):
                os.makedirs(alarm_path)
            wms_mq_queue = deploy().options(self._gra)
            upper_queue = []
            for i in wms_mq_queue:
                upper_queue.append(i.upper())
            if rows:
                ti = time.strftime("%Y-%m-%d")
                if os.path.exists("{}\\pro_queue_{}.txt".format(alarm_path, ti)):
                    with open("{}\\pro_queue_{}.txt".format(alarm_path, ti), 'r') as ff:
                        old_data = eval(ff.read())
                else:
                    old_data = tuple()
                    log().info(
                        '| ALARM CLASS | FORMAT MESS FUNCTION | Message：'
                        'No historical data file for querying MQ exceptions |')
                if not set(rows).issubset(set(old_data)):   # Judge the difference set().issubset()
                    log().info('| ALARM CLASS | FORMAT MESS FUNCTION | Message：Alarm result' + str(rows) + ' |')
                    for tuple1 in rows:
                        for ii in upper_queue:
                            if ii in tuple1[0]:
                                error_queue = ii
                                error_mess = tuple1[0]
                                error_clock = tuple1[1]
                                acknowledged = tuple1[2]
                                mess = tem_mess.format(error_queue, error_mess, error_clock)
                                queue_li = eval(deploy().get(self._gra, error_queue))
                                # acknowledged 0=No 1=Yes
                                if queue_li[0] == 0 and acknowledged == 0 and tuple1 not in old_data:
                                    nickname = queue_li[1]
                                    group_name = queue_li[2]
                                    log().info('| ALARM CLASS | FORMAT MESS FUNCTION | Message：' + mess + ' |')
                                    self.element(url_mq, gra_user, gra_passwd, group_name)
                                    self.alarm_send(nickname, group_name)
                                    self.alarm_send(mess, group_name)
                                    log().info('| ALARM CLASS | FORMAT MESS FUNCTION | Message：'
                                               'The monitoring alarm information is sent successfully.'
                                               'Nickname {} Message {}'.format(nickname, mess))
                                else:
                                    log().info('| ALARM CLASS | FORMAT MESS FUNCTION | Message：'
                                               'ACK == YES, The administrator is aware of the issue |')
                    with open("{}\\pro_queue_{}.txt".format(alarm_path, ti), 'w') as f:
                        f.write(str(rows))
                else:
                    log().info('| ALARM CLASS | FORMAT MESS FUNCTION | Message：'
                               'Zabbix monitoring abnormal data is the same as the historical data, '
                               'and no alarm is required. |')
            else:
                log().info(
                    '| ALARM CLASS | FORMAT MESS FUNCTION | Message：'
                    'Zabbix monitoring did not query the exception queue！|')
        except Exception as e:
            error().error(str(e))
            error().error(traceback.format_exc())

    # QUERY THE ZABBIX PROBLEM VIEW
    def query_alarm(self):
        obj1 = Mysql()
        rows = obj1.execute_sql(self._alarm_sql)
        self.format_mess(rows)


# CONNECT MYSQL METHOD
class Mysql:
    def __init__(self):
        self._sve_file = shelve.open("gywl")
        self._host = self._sve_file['zabbix']['host']
        self._port = self._sve_file['zabbix']['port']
        self._user = self._sve_file['zabbix']['user']
        self._password = self._sve_file['zabbix']['passwd']
        self._db = self._sve_file['zabbix']['db']
        self._charset = self._sve_file['zabbix']['charset']
        self.conn = None
        self.cursor = None
        self.isClose = True

    def connect(self):  # connect
        try:
            self.conn = pymysql.connect(host=self._host,
                                        port=self._port,
                                        user=self._user,
                                        passwd=self._password,
                                        db=self._db,
                                        charset=self._charset)
            self.cursor = self.conn.cursor()
            self.isClose = False
        except Exception as e:
            error().error(
                "| MYSQL CLASS | CONNECT FUNCTION | Message ："
                "Mysql connect error![MSG={},USER={},PASSWORD={},IP={},DB={}]".format(
                    str(e),
                    '***',
                    '***',
                    self._host,
                    self._db))
            error().error(traceback.format_exc())
            obj1 = Comp('@all', traceback.format_exc())
            obj1.post()
            self.cursor = None
            self.isClose = True
        return self.cursor

    def is_connect(self):  # connected or not
        if self.cursor is not None:
            return True
        else:
            return False

    def close(self):  # close oracle connect
        self.cursor.close()
        self.conn.close()
        self.isClose = True

    def execute_sql(self, sql):  # execute fetchall
        if not self.is_connect():
            self.connect()  # True = Not False
            num = 0
            for i in range(3):  # judge 3 time
                num += 1
                if not self.is_connect():
                    self.connect()
                    log().info(
                        '| MYSQL CLASS | EXECUTE SQL FUNCTION | Message：'
                        'Try to reconnect to the database for the {} time'.format(
                            str(num)))
                    time.sleep(1)
                else:
                    break
        else:  # False = Not True
            pass
        try:
            self.cursor.execute(sql)
            if 'declare begin' in sql:
                pass
            else:
                global sql_rows
                sql_rows = self.cursor.fetchall()
        except Exception as e:
            error().error(
                "| MYSQL CLASS | EXECUTE SQL FUNCTION | Message："
                "Execute sql error![MSG={},SQL={}]".format(str(e), sql) + ' |')
            error().error(traceback.format_exc())
            obj1 = Comp('@all', traceback.format_exc())
            obj1.post()
            return ''
        else:
            return sql_rows


def del_link():
    li_folder = eval(deploy().get('folder', 'name'))
    file_limit = eval(deploy().get('file', 'limit'))
    current_path = os.getcwd()
    try:
        for folder in li_folder:
            path = current_path + "\\formatLogs\\{}".format(folder)
            li_file = os.listdir(path)
            count_file = len(li_file)
            if count_file > file_limit:
                for i in range(count_file - file_limit):
                    os.unlink(li_file[i])
                    log().info(
                        '| DEL LINK FUNCTION | Message：Current folder：{}，Delete file：{} |'.format(path, li_file[i]))
            else:
                log().info(
                    '| DEL LINK FUNCTION | Message：'
                    'Current folder：{}，The numbers of file is less than {}，Do not delete file |'.format(
                        path, file_limit))
            li_file_new = os.listdir(path)
            count_file_new = len(li_file_new)
            log().info('| DEL LINK FUNCTION | Message：'
                       'Current folder：{}，The Numbers of file is {}，File {} |'.format(path, count_file_new,
                                                                                      li_file_new))
    except Exception as e:
        error().error(str(e))
        error().error(traceback.format_exc())


def log():
    lg = logformat.Logger(logger=py_name + '[debug]',
                          filename='access',
                          level='debug')  # debug Output to the console | info not
    return lg


def error():
    lg = logformat.Logger(logger=py_name + '[error]',
                          filename='error',
                          level='error')
    return lg


def main():
    global gra_user, gra_passwd, url_mq, tem_mess, alarm_path, py_name, user, password, ip, name, mutex
    file = shelve.open("gywl")
    gra_user = file['grafana']['user']
    gra_passwd = file['grafana']['passwd']
    user = file['devOps']['user']
    password = file['devOps']['passwd']
    ip = file['devOps']['ip']
    name = file['devOps']['name']
    file.close()
    url_mq = deploy().get('ins', 'url_mq')
    tem_mess = deploy().get('ins', 'mess')
    py_name = os.path.basename(sys.argv[0])
    all_time = eval(deploy().get('Timer', 'allTime'))
    kh_time = eval(deploy().get('Timer', 'khTime'))
    erp_time = eval(deploy().get('Timer', 'erpTime'))
    hr_time = eval(deploy().get('Timer', 'hrTime'))
    alarm_time = eval(deploy().get('Timer', 'alarmTime'))
    clean_time = eval(deploy().get('Timer', 'cleanTime'))
    check_time = eval(deploy().get('Timer', 'checkTime'))
    alarm_path = os.getcwd() + "\\formatLogs\\errorQueueFile"
    mutex = Lock()
    order = ORDER()
    file_kh = FILE()
    sh_save = SH()
    check_th = CHECK()
    alarm_q = ALARM()
    order.query_his()
    schedule.every(all_time).minutes.do(order.timer_all)
    schedule.every(kh_time).minutes.do(file_kh.timer_kh)
    schedule.every(erp_time).minutes.do(sh_save.timer_erp)
    schedule.every(hr_time).minutes.do(sh_save.timer_hr)
    schedule.every(alarm_time).minutes.do(alarm_q.query_alarm)
    schedule.every(clean_time).hours.do(clean_screen)
    schedule.every(check_time).minutes.do(check_th.check_thread)
    schedule.every(2).weeks.do(del_link)
    while True:
        schedule.run_pending()
        time.sleep(1)


if __name__ == '__main__':
    print('''
-------------------------------------------------------------------------------------------------
|The task starts executing and is checking whether the current time meets the task requirements.|
|1.The All owner mission is performed from 8:00 to 20:00.                                       |
|2.The HeNan MA  mission is performed from 8:00 to 20:00.                                       |
|3.The KH task first push time on 11:30:00 and the second push time on 16:00:00.                |
|4.The erp|sh|sn mission is performed at 8 o'clock every night.                                 |
|5.The HuiRui mission is performed every hours.                                                 |
|6.The alarm execute every 15 minutes                                                           |
-------------------------------------------------------------------------------------------------''')
    main()
